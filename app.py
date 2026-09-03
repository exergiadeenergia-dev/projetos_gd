"""
Gerador de Documentos de GD — Exergia
App único para gerar os documentos de Geração Distribuída (Energisa-MT e
Equatorial-GO) sem precisar editar JSON ou rodar comandos manualmente.

Rodar localmente:
    pip install -r requirements.txt
    streamlit run app.py

Publicar (acesso pelo navegador de qualquer lugar):
    Veja instruções no LEIA-ME.md
"""
import io
import json
import re
import subprocess
import sys
import tempfile
import zipfile
from pathlib import Path
from datetime import date, datetime, timedelta

import streamlit as st
import pandas as pd
from docx import Document

BASE = Path(__file__).parent
sys.path.insert(0, str(BASE))

from preencher_gd import preencher as preencher_energisa
from recalc import recalc as recalc_xlsx
from office.soffice import run_soffice
from xml_cell_writer import restringir_impressao_para_pdf
from preencher_equatorial import preencher as preencher_anexo1
from preencher_docx_equatorial import (
    replace_everywhere,
    montar_substituicoes_memorial,
    montar_substituicoes_comissionamento,
    fix_tabela_gerador,
    substituir_foto_localizacao,
)
import drive_cliente

PAGINAS_FINAIS_ENERGISA = ["SOLICITACAO", "RELACAO DE CARGA", "FORMULARIO", "MD-SOLAR", "DU-SOLAR"]

# Campos que, quando ficam em branco, a planilha da Energisa destaca em
# vermelho pedindo revisão manual (ex: Potência do Trafo). Sempre que um
# desses campos vier vazio, mostramos um aviso na tela ANTES de gerar os
# documentos, em vez do usuário só descobrir olhando a planilha pronta.
# Para adicionar outro campo a essa checagem, basta incluir a chave (igual
# ao nome usado em dados["cliente"]) e a mensagem de aviso correspondente.
CAMPOS_ATENCAO_ENERGISA = {
    "potencia_trafo": (
        "Potência do Trafo (kVA) não foi informada. A planilha destaca esse "
        "campo em vermelho quando fica em branco — confirme se este projeto "
        "realmente dispensa transformador exclusivo antes de enviar."
    ),
}

# Mesma ideia para o Anexo I da Equatorial-GO. Ainda não temos nenhum campo
# confirmado como "fica vermelho se vazio" nesse template — assim que algum
# for identificado, é só adicionar aqui do mesmo jeito.
CAMPOS_ATENCAO_EQUATORIAL = {}


def avisos_campos_atencao(cliente: dict, campos: dict) -> list:
    """Devolve as mensagens de aviso para os campos de `campos` que estão
    vazios em `cliente` (None, string vazia ou só espaços)."""
    avisos = []
    for chave, mensagem in campos.items():
        valor = cliente.get(chave)
        if valor is None or str(valor).strip() == "":
            avisos.append(mensagem)
    return avisos


def zip_de_arquivos(itens: list) -> bytes:
    """Monta um .zip em memória a partir de uma lista de (nome, bytes)."""
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as z:
        for nome, conteudo in itens:
            z.writestr(nome, conteudo)
    return buffer.getvalue()


def _num(defaults: dict, campo: str, padrao):
    """Lê um valor numérico de `defaults` (que pode vir como string, já
    que o quadro de conferência do Drive edita tudo como texto) devolvendo
    `padrao` (int ou float, conforme o tipo de `padrao`) se não der pra
    converter. Ao contrário do `_txt` acima, aqui `padrao` continua valendo
    mesmo com `defaults` não-vazio — a maioria desses campos (disjuntor, DPS,
    fator de potência, nº de hastes) são valores padrão de engenharia
    reaproveitados projeto a projeto, não dados específicos do cliente, e já
    eram assim antes da busca no Drive existir. Sempre revise antes de gerar."""
    valor = defaults.get(campo)
    if valor is None or str(valor).strip() == "":
        return padrao
    try:
        convertido = float(str(valor).replace(",", "."))
        return int(convertido) if isinstance(padrao, int) else convertido
    except ValueError:
        return padrao


def _txt(defaults: dict, campo: str, padrao: str = "") -> str:
    """Valor de texto pra pré-preencher um campo do formulário manual.
    Se o campo veio do levantamento do Drive/texto colado, usa esse valor.
    Se `defaults` está vazio (formulário em branco, ninguém usou Drive/colar
    ainda), usa o valor de exemplo `padrao` — igual sempre foi.
    Mas se `defaults` NÃO está vazio (ou seja, veio do Drive) e esse campo
    específico não foi encontrado, o campo fica em branco em vez de mostrar
    o valor de exemplo — pra não passar a impressão de que aquele dado foi
    conferido quando na verdade ninguém olhou pra ele."""
    if campo in defaults:
        return str(defaults[campo])
    return padrao if not defaults else ""


def caixa_fonte_inmetro():
    """Lembrete + links oficiais do INMETRO pra achar marca/modelo exatos
    de módulo e inversor. O app nunca escolhe um modelo sozinho — o
    engenheiro sempre confirma o modelo certificado antes de gerar os
    documentos; a única automação aqui é apontar pra fonte certa."""
    st.caption(
        "📋 Marca e modelo de módulo/inversor **sempre são conferidos e confirmados por você** "
        "— o app não escolhe isso sozinho. Use as tabelas oficiais do INMETRO como fonte para "
        f"achar o modelo certificado a partir da potência: [tabela de módulos fotovoltaicos]({drive_cliente.INMETRO_MODULOS_URL}) · "
        f"[tabela de inversores on-grid]({drive_cliente.INMETRO_INVERSORES_ONGRID_URL})."
    )


def _idx(opcoes: list, valor, padrao: int = 0) -> int:
    """Índice de `valor` dentro de `opcoes` pra pré-selecionar um
    selectbox a partir de um dado vindo do levantamento automático do
    Drive; devolve `padrao` se o valor não bater com nenhuma opção."""
    if valor is None:
        return padrao
    valor_norm = str(valor).strip().upper()
    for i, opcao in enumerate(opcoes):
        if str(opcao).strip().upper() == valor_norm:
            return i
    return padrao


LOGO_PATH = BASE / "logo_exergia.png"

st.set_page_config(
    page_title="Gerador GD — Exergia",
    page_icon=str(LOGO_PATH) if LOGO_PATH.exists() else "☀️",
    layout="wide",
)
if LOGO_PATH.exists():
    st.logo(str(LOGO_PATH), size="large")

MODELO_ENERGISA = BASE / "memorialgd_1.xlsm"
MODELO_ANEXO1 = BASE / "NT_00020_EQTL-06-Anexo-I-Formulario-de-Solicitacao.xlsx"
MODELO_MEMORIAL = BASE / "Memorial_Descritivo_MODELO.docx"
MODELO_COMISSIONAMENTO = BASE / "Modelo_Comissionamento_MODELO.docx"

REFERENCIA_HAKKNNER = {
    "nome": "Fulano de Tal da Silva",
    "nome_maiusculo": "FULANO DE TAL DA SILVA",
    "conta_contrato": "0.000.000.000-00",
    "endereco_uc": "Rua Exemplo, Q. 00 L. 00 S/N.",
    "bairro_cidade": "Setor Exemplo. Aragarças – GO.",
    "cep": "76240000",
    "cidade": "Aragarças",
    "coordenadas": "UTM 22K, X: 366859 Y: 8240498",
    "equip_titulo": "5 kW",
    "equip_desc": "8 módulos OSDA/ODA590-36V-MHD",
    "modalidade": "autoconsumo remoto",
    "mes_ano": "Junho – 2026",
    "estado": "Goiás",
    "fab_modulo": "OSDA",
    "modelo_modulo": "ODA590-36V-MHD",
    "telefone_cliente": "(00) 00000-0000",
    "uc_pontuado": "0.000.000.000-00",
    "data_conclusao": "24/06/2026",
    "endereco_completo": "RUA EXEMPLO, Q. 00, L. 00, S/N, CASA - RESIDENCIA\nSETOR EXEMPLO",
    "fab_inversor": "SOFAR",
    "modelo_inversor": "5KTLM-G3",
    "fab_inversor_typo_memorial": "SOAR",
}

MES_PT = ["JANEIRO", "FEVEREIRO", "MARÇO", "ABRIL", "MAIO", "JUNHO", "JULHO",
          "AGOSTO", "SETEMBRO", "OUTUBRO", "NOVEMBRO", "DEZEMBRO"]


def slug_arquivo(nome: str, uc) -> str:
    base = str(nome).strip().upper().replace(" ", "_")
    uc_limpo = str(uc).strip().replace(" ", "_")
    return f"{base}_UC_{uc_limpo}"


def responsavel_tecnico_form():
    with st.expander("Responsável técnico (raramente muda)"):
        c1, c2 = st.columns(2)
        nome = c1.text_input("Nome", "Alex Gomes da Silva")
        email = c2.text_input("E-mail", "alexgomes.eng.energ@gmail.com")
        c3, c4 = st.columns(2)
        telefone = c3.text_input("Telefone", "(66) 9 9964-5914")
        registro = c4.text_input("Registro profissional (CREA)", "043613")
    return {"nome": nome, "email": email, "telefone": telefone, "registro": registro}


def tabela_equipamentos(label, colunas, valores_padrao, key):
    st.caption(label)
    df = pd.DataFrame(valores_padrao)
    edited = st.data_editor(df, num_rows="dynamic", key=key, width="stretch")
    return edited.to_dict("records")


MESES_EN_PT = {
    1: "JANEIRO", 2: "FEVEREIRO", 3: "MARÇO", 4: "ABRIL", 5: "MAIO", 6: "JUNHO",
    7: "JULHO", 8: "AGOSTO", 9: "SETEMBRO", 10: "OUTUBRO", 11: "NOVEMBRO", 12: "DEZEMBRO",
}


def _converter_valor(v: str):
    """Tenta converter para int, depois float; senão devolve a string original."""
    v = v.strip()
    try:
        return int(v)
    except ValueError:
        pass
    try:
        return float(v.replace(",", "."))
    except ValueError:
        pass
    return v


def parse_texto_colado(texto: str) -> dict:
    """Faz o parse do formato 'chave: valor' com blocos repetíveis [nome]
    (ex: [cargas], [paineis], [inversores]) para um dict pronto pra gerar
    os documentos, sem precisar preencher tabela por tabela na tela."""
    campos = {}
    blocos = {}  # nome_do_bloco -> lista de dicts
    bloco_atual = None
    item_atual = None

    for linha_bruta in texto.splitlines():
        linha = linha_bruta.strip()
        if not linha:
            continue
        if linha.startswith("----") or linha.startswith("#"):
            # uma linha separadora de seção sempre fecha qualquer bloco
            # [cargas]/[paineis]/[inversores] que ainda esteja aberto —
            # sem isso, os campos da seção seguinte (ex: disjuntor_geral_a)
            # acabavam grudados dentro do último item do bloco anterior.
            if bloco_atual and item_atual is not None:
                blocos.setdefault(bloco_atual, []).append(item_atual)
                bloco_atual = None
                item_atual = None
            continue
        m = re.match(r"^\[(\w+)\]$", linha)
        if m:
            if bloco_atual and item_atual is not None:
                blocos.setdefault(bloco_atual, []).append(item_atual)
            bloco_atual = m.group(1).lower()
            item_atual = {}
            continue
        if ":" in linha:
            chave, valor = linha.split(":", 1)
            chave = chave.strip().lower()
            valor = _converter_valor(valor)
            if item_atual is not None:
                item_atual[chave] = valor
            else:
                campos[chave] = valor

    if bloco_atual and item_atual is not None:
        blocos.setdefault(bloco_atual, []).append(item_atual)

    campos["_blocos"] = blocos

    # normaliza data de previsão de ligação (aceita "Sep 15, 2026",
    # "15/09/2026", "2026-09-15" etc.) para mês/ano em português
    data_str = campos.pop("data_previsao_ligacao", None)
    if data_str:
        data_convertida = None
        for fmt in ("%b %d, %Y", "%B %d, %Y", "%d/%m/%Y", "%Y-%m-%d"):
            try:
                data_convertida = datetime.strptime(str(data_str).strip(), fmt)
                break
            except ValueError:
                continue
        if data_convertida:
            campos["mes_previsao_ligacao"] = MESES_EN_PT[data_convertida.month]
            campos["ano_previsao_ligacao"] = data_convertida.year

    return campos


def campo_colar_dados(exemplo: str, ajuda_blocos: str) -> dict | None:
    """Mostra a caixa de texto pra colar os dados e devolve o dict parseado,
    ou None se o usuário não colou nada."""
    with st.expander("📋 Colar dados em texto (mais rápido que preencher campo por campo)"):
        st.caption(
            "Cole os dados no formato `chave: valor`, um por linha. "
            f"Use blocos repetíveis pra listas: {ajuda_blocos}"
        )
        caixa_fonte_inmetro()
        st.code(exemplo, language="text")
        texto = st.text_area("Cole aqui", height=200, key="colar_" + ajuda_blocos[:10])
    if texto and texto.strip():
        return parse_texto_colado(texto)
    return None


# ============================================================
# GOOGLE DRIVE — levantamento automático de dados
# ============================================================
NOME_TIPO = {"energisa": "Energisa-MT", "equatorial": "Equatorial-GO"}

# de campo do levantamento (dict `campos` do drive_cliente) -> chave usada
# nos widgets do formulário manual de cada concessionária. Onde o nome é
# igual dos dois lados a entrada nem precisa aparecer aqui.
MAPA_CAMPOS_EQUATORIAL = {
    "titular": "nome",
    "logradouro": "endereco",
    "cidade": "municipio",
    "tipo_conexao": "tipo_ligacao",
}


def _remapear_para_equatorial(campos: dict) -> dict:
    """Aplica MAPA_CAMPOS_EQUATORIAL e junta logradouro+número num só campo
    "endereco" (o formulário da Equatorial-GO não tem campo separado pra
    número — é "Endereço (rua, número, complemento)" tudo numa caixa só)."""
    campos = dict(campos)
    numero = campos.pop("numero", None)
    if numero and campos.get("logradouro"):
        campos["logradouro"] = f"{campos['logradouro']}, {numero}"
    return {MAPA_CAMPOS_EQUATORIAL.get(c, c): v for c, v in campos.items()}


def _guia_configuracao_drive():
    st.info(
        "🔒 A leitura automática do Drive ainda não está configurada neste app. "
        "Isso é feito uma única vez (veja o passo a passo abaixo) — até lá, use "
        "normalmente a opção de colar os dados em texto ou o formulário manual."
    )
    with st.expander("Como configurar (uma vez só)"):
        st.markdown(
            "1. No [Google Cloud Console](https://console.cloud.google.com/), crie um "
            "projeto (ou use um existente) e ative a **Google Drive API**.\n"
            "2. Em *Credenciais*, crie uma **Conta de serviço** e gere uma **chave JSON**.\n"
            "3. No painel do Streamlit Cloud, vá em *Settings → Secrets* e cole o "
            "conteúdo do JSON sob a chave `[gcp_service_account]` (no formato TOML).\n"
            "4. No Google Drive, compartilhe a pasta-mãe dos clientes (somente leitura) "
            "com o e-mail da conta de serviço — algo como "
            "`nome@seu-projeto.iam.gserviceaccount.com` (está dentro do JSON, campo "
            "`client_email`).\n\n"
            "A partir daí o app só enxerga as pastas que forem compartilhadas com essa "
            "conta — nada além disso, e nunca escreve nada no Drive."
        )


def painel_busca_drive(tipo: str) -> dict:
    """Mostra a busca da pasta do cliente no Drive + o levantamento
    automático dos dados. Assim que o levantamento roda, os campos
    encontrados já descem direto pro formulário manual logo abaixo — não
    tem uma segunda etapa de confirmação aqui; a conferência e a edição
    acontecem direto nos campos do formulário (que é a própria "tela de
    confirmar"), e o botão de gerar já existente lá embaixo é o único
    passo final. Devolve um dict de valores pra pré-preencher esse
    formulário (vazio se o usuário não usou o Drive nesta sessão)."""
    chave_resultado = f"drive_resultado_{tipo}"
    chave_pastas = f"drive_pastas_{tipo}"
    chave_confirmado = f"drive_confirmado_{tipo}"

    with st.expander(f"🔍 Buscar dados automaticamente no Google Drive ({NOME_TIPO[tipo]})"):
        if not drive_cliente.servico_disponivel():
            _guia_configuracao_drive()
            return st.session_state.get(chave_confirmado, {})

        st.caption(
            "Digite o nome do cliente (busca pelas pastas compartilhadas com o app) ou "
            "cole o link da pasta do Drive. O app só lê os arquivos dessa pasta — nada "
            "é escrito, editado ou apagado no Drive."
        )
        busca = st.text_input("Nome do cliente ou link/ID da pasta", key=f"drive_busca_{tipo}")

        if st.button("Buscar no Drive", key=f"drive_btn_buscar_{tipo}"):
            try:
                servico = drive_cliente._get_service()
                if "/" in busca or len(busca) > 40:
                    pasta_id = drive_cliente.extrair_id_pasta(busca)
                    meta = servico.files().get(fileId=pasta_id, fields="id, name").execute()
                    st.session_state[chave_pastas] = [{"id": meta["id"], "name": meta["name"]}]
                else:
                    st.session_state[chave_pastas] = drive_cliente.buscar_pastas(servico, busca)
                if not st.session_state[chave_pastas]:
                    st.warning("Nenhuma pasta encontrada com esse nome/link.")
            except Exception as exc:
                st.error(f"Não consegui acessar o Drive: {exc}")

        pastas = st.session_state.get(chave_pastas, [])
        if pastas:
            opcoes = {f"{p['name']} (id: {p['id'][:8]}…)": p["id"] for p in pastas}
            escolha = st.radio("Pasta encontrada", list(opcoes.keys()), key=f"drive_radio_{tipo}")
            pasta_id = opcoes[escolha]

            if st.button(f"📥 Fazer levantamento automático — {NOME_TIPO[tipo]}", key=f"drive_btn_levantar_{tipo}"):
                with st.spinner("Lendo os arquivos da pasta e reconhecendo os dados..."):
                    try:
                        resultado = drive_cliente.levantamento_pasta(pasta_id, tipo)
                        st.session_state[chave_resultado] = resultado
                        # já aplica direto no formulário manual abaixo — sem
                        # etapa extra de confirmação aqui em cima.
                        campos = dict(resultado.campos)
                        if tipo == "equatorial":
                            campos = _remapear_para_equatorial(campos)
                        st.session_state[chave_confirmado] = campos
                    except Exception as exc:
                        st.error(f"Erro durante o levantamento: {exc}")

        resultado = st.session_state.get(chave_resultado)
        if resultado:
            if resultado.arquivos_lidos:
                st.success("Arquivos lidos: " + ", ".join(resultado.arquivos_lidos))
            if resultado.arquivos_ignorados:
                st.caption("Ignorados (sem texto legível ou tipo não suportado): "
                            + ", ".join(resultado.arquivos_ignorados))
            for erro in resultado.erros:
                st.warning(erro)

            if not resultado.campos:
                st.warning("Não consegui reconhecer nenhum dado nos arquivos dessa pasta. "
                           "Preencha manualmente abaixo ou use a opção de colar texto.")
            else:
                linhas_fonte = "\n".join(
                    f"- **{campo}**: {valor}  ·  fonte: {resultado.fontes.get(campo, '')}"
                    for campo, valor in sorted(resultado.campos.items())
                )
                st.success(
                    "Campos já aplicados no formulário abaixo — role até lá pra conferir e "
                    "corrigir cada um antes de clicar em Gerar documentos:\n\n" + linhas_fonte
                )

        return st.session_state.get(chave_confirmado, {})


# ============================================================
# ENERGISA-MT
# ============================================================
def form_energisa():
    st.header("Energisa-MT")
    mostrar_downloads_energisa()

    defaults = painel_busca_drive("energisa")

    exemplo = """uc: 74561701723
classe: RESIDENCIAL
titular: Nome Completo do Cliente
logradouro: Rua Exemplo
numero: 100
bairro: Centro
cidade: Barra do Garças
uf: MT
cep: 78600-000
email: cliente@email.com
celular: 66 99999-9999
cpf_cnpj: 000.000.000-00
potencia_instalada_kw: 15
tensao_atendimento_v: 220
tipo_conexao: MONOFÁSICO
tipo_ramal: AÉREO

---- Relação de carga: repita o bloco [cargas] para cada equipamento ----
[cargas]
quantidade: 2
equipamento: CHUVEIRO
pot_unitaria_w: 4500
fator_demanda: 0.82

---- Proteções e padrão ----
disjuntor_geral_a: 50
fator_potencia: 0.92
demanda_contratada_kw: 0
dps_ca_ka: 20
disjuntor_ca_a: 50
dps_cc_ka: 20
disjuntor_cc_a: 32
modalidade: Autoconsumo remoto
potencia_trafo: 45
numero_hastes: 6

---- Coordenadas e previsão de ligação ----
fuso: 22L
coord_x: 358297
coord_y: 8245026
data_previsao_ligacao: Sep 15, 2026
zona: URBANO
gd_ja_instalado: NÃO

---- Painéis: repita o bloco [paineis] para cada modelo ----
[paineis]
quantidade: 9
fabricante: Maxeon
modelo: SPR-P6-550-UPP
area_m2: 18
potencia_kw: 0.55

---- Inversores: repita o bloco [inversores] para cada modelo ----
[inversores]
quantidade: 1
fabricante: Auxsol
modelo: ASN-5SL-G2
potencia_kw: 5
tensao_nominal_v: 220"""

    colado = campo_colar_dados(exemplo, "[cargas], [paineis], [inversores]")

    if colado:
        st.success("Texto reconhecido! Confira o resumo abaixo antes de gerar.")
        blocos = colado.pop("_blocos", {})
        cargas = blocos.get("cargas", [])
        paineis = blocos.get("paineis", [])
        inversores = blocos.get("inversores", [])

        faltando = [c for c in ("uc", "titular") if not colado.get(c)]
        st.write(
            f"**{colado.get('titular','?')}** — UC {colado.get('uc','?')} — "
            f"{len(cargas)} carga(s), {len(paineis)} modelo(s) de painel, "
            f"{len(inversores)} modelo(s) de inversor"
        )
        with st.expander("Ver dados completos reconhecidos"):
            st.json({**colado, "cargas": cargas, "paineis": paineis, "inversores": inversores})

        if faltando:
            st.error(f"Faltam campos obrigatórios no texto colado: {', '.join(faltando)}")
            return
        if not cargas:
            st.warning("Nenhuma carga (bloco [cargas]) reconhecida no texto.")
        if not paineis:
            st.warning("Nenhum painel (bloco [paineis]) reconhecido no texto.")
        if not inversores:
            st.warning("Nenhum inversor (bloco [inversores]) reconhecido no texto.")

        if st.button("Gerar documentos a partir do texto colado (Energisa-MT)", type="primary"):
            dados = {
                "responsavel_tecnico": {
                    "nome": "Alex Gomes da Silva",
                    "telefone": "(66) 9 9964-5914",
                    "email": "alexgomes.eng.energ@gmail.com",
                },
                "cliente": {
                    "uc": colado.get("uc", ""), "classe": colado.get("classe", "RESIDENCIAL"),
                    "titular": colado.get("titular", ""), "logradouro": colado.get("logradouro", ""),
                    "numero": colado.get("numero", ""), "bairro": colado.get("bairro", ""),
                    "cidade": colado.get("cidade", ""), "uf": colado.get("uf", "MT"),
                    "cep": colado.get("cep", ""), "email": colado.get("email", "não informado"),
                    "telefone": "", "celular": colado.get("celular", ""),
                    "cpf_cnpj": colado.get("cpf_cnpj", ""),
                    "potencia_instalada_kw": colado.get("potencia_instalada_kw", 0),
                    "tensao_atendimento_v": str(colado.get("tensao_atendimento_v", "220")),
                    "tipo_conexao": colado.get("tipo_conexao", "MONOFÁSICO"),
                    "tipo_ramal": colado.get("tipo_ramal", "AÉREO"),
                    "tipo_fonte_geracao": "SOLAR FOTOVOLTAICA",
                    "tipo_geracao": "Empregando conversor eletrônico/inversor",
                    "cargas": cargas,
                    "disjuntor_geral_a": colado.get("disjuntor_geral_a", 50),
                    "fator_potencia": colado.get("fator_potencia", 0.92),
                    "demanda_contratada_kw": colado.get("demanda_contratada_kw", 0),
                    "dps_ca_ka": colado.get("dps_ca_ka", 20),
                    "disjuntor_ca_a": colado.get("disjuntor_ca_a", 50),
                    "dps_cc_ka": colado.get("dps_cc_ka", 20),
                    "disjuntor_cc_a": colado.get("disjuntor_cc_a", 32),
                    "modalidade": colado.get("modalidade", "Autoconsumo remoto"),
                    "potencia_trafo": colado.get("potencia_trafo", ""),
                    "numero_hastes": colado.get("numero_hastes", 6),
                    "demanda_contratada_kwg": 0,
                    "fuso": colado.get("fuso", ""), "coord_x": colado.get("coord_x", ""),
                    "coord_y": colado.get("coord_y", ""),
                    "tipo_tensao": "BAIXA", "cabos_por_fase": 1,
                    "bitola_fase": 10, "bitola_neutro": 10, "bitola_terra": 10,
                    "gd_ja_instalado": colado.get("gd_ja_instalado", "SIM"),
                    "mes_previsao_ligacao": colado.get("mes_previsao_ligacao", MES_PT[date.today().month - 1]),
                    "ano_previsao_ligacao": colado.get("ano_previsao_ligacao", date.today().year),
                    "zona": colado.get("zona", "URBANO"),
                    "paineis": paineis, "inversores": inversores,
                    "necessita_autotrafo": "NÃO", "trafo_exclusivo": "NÃO",
                },
            }
            for aviso in avisos_campos_atencao(dados["cliente"], CAMPOS_ATENCAO_ENERGISA):
                st.warning(f"⚠️ {aviso}")
            gerar_energisa(dados)
        return

    st.divider()
    if defaults:
        st.caption("💡 Os campos abaixo já vêm preenchidos com os dados confirmados no "
                   "quadro do Drive acima — revise mais uma vez antes de gerar.")
    resp = responsavel_tecnico_form()

    st.subheader("1. Dados da Unidade Consumidora")
    c1, c2, c3 = st.columns(3)
    uc = c1.text_input("UC", defaults.get("uc", ""))
    classe = c2.selectbox("Classe", ["RESIDENCIAL", "COMERCIAL", "INDUSTRIAL", "RURAL"],
                           index=_idx(["RESIDENCIAL", "COMERCIAL", "INDUSTRIAL", "RURAL"], defaults.get("classe")))
    titular = c3.text_input("Titular", defaults.get("titular", ""))

    c1, c2, c3 = st.columns(3)
    logradouro = c1.text_input("Logradouro", defaults.get("logradouro", ""))
    numero = c2.text_input("Número", defaults.get("numero", ""))
    bairro = c3.text_input("Bairro", defaults.get("bairro", ""))

    c1, c2, c3 = st.columns(3)
    cidade = c1.text_input("Cidade", defaults.get("cidade", ""))
    uf = c2.text_input("UF", _txt(defaults, "uf", "MT"))
    cep = c3.text_input("CEP", defaults.get("cep", ""))

    c1, c2, c3 = st.columns(3)
    email = c1.text_input("E-mail", defaults.get("email", "não informado"))
    celular = c2.text_input("Celular", defaults.get("celular", ""))
    cpf_cnpj = c3.text_input("CPF/CNPJ", defaults.get("cpf_cnpj", ""))

    st.subheader("2. Dados da UC no ato da vistoria")
    c1, c2, c3, c4 = st.columns(4)
    potencia_instalada_kw = c1.number_input("Potência Instalada (kW)", min_value=0.0, step=0.1,
                                             value=_num(defaults, "potencia_instalada_kw", 0.0))
    tensao_atendimento_v = c2.text_input("Tensão de Atendimento (V)",
                                          _txt(defaults, "tensao_atendimento_v", "220"),
                                          help='Use "220/380" para trifásico em baixa tensão')
    opcoes_conexao = ["MONOFÁSICO", "BIFÁSICO", "TRIFÁSICO"]
    tipo_conexao = c3.selectbox("Tipo de Conexão", opcoes_conexao,
                                 index=_idx(opcoes_conexao, defaults.get("tipo_conexao")))
    opcoes_ramal = ["AÉREO", "SUBTERRÂNEO"]
    tipo_ramal = c4.selectbox("Tipo de Ramal", opcoes_ramal,
                               index=_idx(opcoes_ramal, defaults.get("tipo_ramal")))

    st.subheader("3. Relação de Carga")
    cargas = tabela_equipamentos(
        "Equipamentos (adicione/remova linhas como precisar)",
        None,
        [
            {"quantidade": 2, "equipamento": "CHUVEIRO", "pot_unitaria_w": 4500, "fator_demanda": 0.82},
            {"quantidade": 2, "equipamento": "Ar Condicionado", "pot_unitaria_w": 1300, "fator_demanda": 0.82},
            {"quantidade": 1, "equipamento": "Eletrodomésticos", "pot_unitaria_w": 2000, "fator_demanda": 0.8},
            {"quantidade": 1, "equipamento": "Eletrônicos", "pot_unitaria_w": 800, "fator_demanda": 0.8},
        ],
        "cargas_energisa",
    )

    st.subheader("4. Proteções e padrão")
    c1, c2, c3 = st.columns(3)
    disjuntor_geral_a = c1.number_input("Disjuntor geral (A)", value=_num(defaults, "disjuntor_geral_a", 50))
    fator_potencia = c2.number_input("Fator de Potência", value=_num(defaults, "fator_potencia", 0.92), step=0.01)
    demanda_contratada_kw = c3.number_input("Demanda Contratada (kW)", value=_num(defaults, "demanda_contratada_kw", 0.0))

    c1, c2, c3, c4 = st.columns(4)
    dps_ca_ka = c1.number_input("DPS CA (kA)", value=_num(defaults, "dps_ca_ka", 20))
    disjuntor_ca_a = c2.number_input("Disjuntor CA (A)", value=_num(defaults, "disjuntor_ca_a", 50))
    dps_cc_ka = c3.number_input("DPS CC (kA)", value=_num(defaults, "dps_cc_ka", 20))
    disjuntor_cc_a = c4.number_input("Disjuntor CC (A)", value=_num(defaults, "disjuntor_cc_a", 32))

    c1, c2, c3 = st.columns(3)
    opcoes_modalidade = ["Autoconsumo remoto", "Autoconsumo local", "Geração compartilhada", "Múltiplas UCs"]
    modalidade = c1.selectbox("Modalidade", opcoes_modalidade,
                               index=_idx(opcoes_modalidade, defaults.get("modalidade")))
    potencia_trafo = c2.text_input("Potência Trafo (kVA) — deixe em branco se não houver",
                                    defaults.get("potencia_trafo", ""))
    numero_hastes = c3.number_input("Número de hastes", value=_num(defaults, "numero_hastes", 6))

    st.subheader("5. Coordenadas e previsão de ligação")
    c1, c2, c3 = st.columns(3)
    fuso = c1.text_input("Fuso (ex: 22L)", defaults.get("fuso", ""))
    coord_x = c2.text_input("X", defaults.get("coord_x", ""))
    coord_y = c3.text_input("Y", defaults.get("coord_y", ""))

    c1, c2, c3 = st.columns(3)
    data_prevista = c1.date_input("Previsão de ligação", value=date.today() + timedelta(days=15))
    opcoes_zona = ["URBANO", "RURAL"]
    zona = c2.selectbox("Zona", opcoes_zona, index=_idx(opcoes_zona, defaults.get("zona")))
    gd_ja_instalado = c3.selectbox("Sistema GD já instalado?", ["SIM", "NÃO"])

    st.subheader("6. Painéis")
    caixa_fonte_inmetro()
    paineis = tabela_equipamentos(
        "Um item por modelo de painel diferente",
        None,
        [{"quantidade": 0, "fabricante": "", "modelo": "", "area_m2": 0, "potencia_kw": 0.0}],
        "paineis_energisa",
    )

    st.subheader("7. Inversores")
    inversores = tabela_equipamentos(
        "Um item por modelo de inversor diferente",
        None,
        [{"quantidade": 0, "fabricante": "", "modelo": "", "potencia_kw": 0.0, "tensao_nominal_v": 220}],
        "inversores_energisa",
    )

    gerar = st.button("Gerar documentos (Energisa-MT)", type="primary")

    if gerar:
        if not uc or not titular:
            st.error("Preencha ao menos UC e Titular.")
            return

        dados = {
            "responsavel_tecnico": {
                "nome": resp["nome"], "telefone": resp["telefone"], "email": resp["email"],
            },
            "cliente": {
                "uc": uc, "classe": classe, "titular": titular, "logradouro": logradouro,
                "numero": numero, "bairro": bairro, "cidade": cidade, "uf": uf, "cep": cep,
                "email": email, "telefone": "", "celular": celular, "cpf_cnpj": cpf_cnpj,
                "potencia_instalada_kw": potencia_instalada_kw,
                "tensao_atendimento_v": tensao_atendimento_v,
                "tipo_conexao": tipo_conexao, "tipo_ramal": tipo_ramal,
                "tipo_fonte_geracao": "SOLAR FOTOVOLTAICA",
                "tipo_geracao": "Empregando conversor eletrônico/inversor",
                "cargas": cargas,
                "disjuntor_geral_a": disjuntor_geral_a, "fator_potencia": fator_potencia,
                "demanda_contratada_kw": demanda_contratada_kw,
                "dps_ca_ka": dps_ca_ka, "disjuntor_ca_a": disjuntor_ca_a,
                "dps_cc_ka": dps_cc_ka, "disjuntor_cc_a": disjuntor_cc_a,
                "modalidade": modalidade,
                "potencia_trafo": potencia_trafo, "numero_hastes": numero_hastes,
                "demanda_contratada_kwg": 0,
                "fuso": fuso, "coord_x": coord_x, "coord_y": coord_y,
                "tipo_tensao": "BAIXA", "cabos_por_fase": 1,
                "bitola_fase": 10, "bitola_neutro": 10, "bitola_terra": 10,
                "gd_ja_instalado": gd_ja_instalado,
                "mes_previsao_ligacao": MES_PT[data_prevista.month - 1],
                "ano_previsao_ligacao": data_prevista.year,
                "zona": zona,
                "paineis": paineis, "inversores": inversores,
                "necessita_autotrafo": "NÃO", "trafo_exclusivo": "NÃO",
            },
        }
        for aviso in avisos_campos_atencao(dados["cliente"], CAMPOS_ATENCAO_ENERGISA):
            st.warning(f"⚠️ {aviso}")
        gerar_energisa(dados)


def gerar_energisa(dados):
    nome_base = slug_arquivo(dados["cliente"]["titular"], dados["cliente"]["uc"])
    with tempfile.TemporaryDirectory() as tmp:
        tmp = Path(tmp)
        xlsm_path = tmp / f"{nome_base}.xlsm"

        with st.spinner("Preenchendo planilha..."):
            preencher_energisa(dados, str(MODELO_ENERGISA), str(xlsm_path))

        with st.spinner("Recalculando fórmulas..."):
            resultado = recalc_xlsx(str(xlsm_path), 90)
        if resultado.get("total_errors", 0) > 1 or (
            resultado.get("total_errors") == 1
            and "CONFIG!Q27" not in str(resultado.get("error_summary", {}))
        ):
            st.warning(f"Atenção: erros encontrados na planilha: {resultado}")

        with st.spinner("Gerando PDF..."):
            # Cópia só pra imprimir: restringe a área de impressão das demais
            # abas por edição direta do XML (preserva as formas do diagrama —
            # NUNCA usar openpyxl aqui, ele apaga os desenhos ao salvar).
            print_path = tmp / f"{nome_base}_print.xlsm"
            import shutil as _shutil
            _shutil.copyfile(str(xlsm_path), str(print_path))
            restringir_impressao_para_pdf(str(print_path), PAGINAS_FINAIS_ENERGISA)

            run_soffice(["--headless", "--convert-to", "pdf", "--outdir", str(tmp), str(print_path)])
            from pypdf import PdfReader, PdfWriter
            gerado_pdf = print_path.with_suffix(".pdf")
            reader = PdfReader(str(gerado_pdf))
            writer = PdfWriter()
            for page in reader.pages:
                if page.extract_text().strip():
                    writer.add_page(page)
            pdf_final = tmp / f"{nome_base}.pdf"
            with open(pdf_final, "wb") as f:
                writer.write(f)

        with st.spinner("Gerando arquivo para ANEEL..."):
            import openpyxl
            wb2 = openpyxl.load_workbook(str(xlsm_path), data_only=True)
            ws_src = wb2["SAIDA"]
            wb_out = openpyxl.Workbook()
            ws_out = wb_out.active
            ws_out.title = "SAIDA"
            for row in ws_src.iter_rows(min_row=1, max_row=2, max_col=ws_src.max_column):
                for cell in row:
                    if cell.value is not None:
                        ws_out.cell(row=cell.row, column=cell.column, value=cell.value)
            wb_out.create_sheet("Planilha1")
            aneel_path = tmp / f"{nome_base}_ANEEL.xlsx"
            wb_out.save(str(aneel_path))

        # Guarda os bytes na sessão (não nos caminhos!) — a pasta temporária
        # é apagada assim que sairmos deste bloco, e clicar num botão de
        # download recarrega a página. Sem isso, os botões (e os arquivos)
        # somem depois do primeiro clique.
        st.session_state["saida_energisa"] = {
            "xlsm_bytes": xlsm_path.read_bytes(), "xlsm_nome": xlsm_path.name,
            "pdf_bytes": pdf_final.read_bytes(), "pdf_nome": pdf_final.name,
            "aneel_bytes": aneel_path.read_bytes(), "aneel_nome": aneel_path.name,
        }

    mostrar_downloads_energisa()


def mostrar_downloads_energisa():
    saida = st.session_state.get("saida_energisa")
    if not saida:
        return
    st.success("Documentos gerados!")
    c1, c2, c3, c4 = st.columns(4)
    c1.download_button("⬇️ Planilha (.xlsm)", saida["xlsm_bytes"], saida["xlsm_nome"], key="dl_xlsm_energisa")
    c2.download_button("⬇️ PDF final", saida["pdf_bytes"], saida["pdf_nome"], key="dl_pdf_energisa")
    c3.download_button("⬇️ Arquivo ANEEL", saida["aneel_bytes"], saida["aneel_nome"], key="dl_aneel_energisa")
    zip_bytes = zip_de_arquivos([
        (saida["xlsm_nome"], saida["xlsm_bytes"]),
        (saida["pdf_nome"], saida["pdf_bytes"]),
        (saida["aneel_nome"], saida["aneel_bytes"]),
    ])
    c4.download_button("📦 Baixar tudo (.zip)", zip_bytes, "documentos_energisa.zip",
                        key="dl_zip_energisa", type="primary")


# ============================================================
# EQUATORIAL-GO
# ============================================================
def form_equatorial():
    st.header("Equatorial-GO")
    mostrar_downloads_equatorial()

    defaults = painel_busca_drive("equatorial")

    exemplo_go = """nome: Nome Completo do Cliente
cpf_cnpj: 000.000.000-00
celular: 66 99999-9999
endereco: Rua Exemplo, 100
email: cliente@email.com
cep: 76240000
municipio: Aragarças
uf: GO
uc_existente: 0000000000
tipo_ligacao: MONOFÁSICO
tensao_atendimento_v: 220
classe: Residencial
disjuntor_entrada_a: 50
carga_declarada_kw: 10
potencia_disponibilizada_kw: 10
tipo_ramal: AÉREO
fuso: 22L
coord_x: 366000
coord_y: 8241000
modalidade_compensacao: AUTOCONSUMO LOCAL
data_inicio_operacao: Sep 15, 2026
bairro: Centro

---- Módulos: repita o bloco [modulos] para cada modelo ----
[modulos]
potencia_w: 590
quantidade: 10
fabricante: LEAPTON
modelo: LP182-199-M-66-NB-590W

---- Inversores: repita o bloco [inversores] para cada modelo ----
[inversores]
fabricante: SOFAR
modelo: 5KTLM-G3
potencia_nominal_kw: 5
faixa_tensao_v: 220
corrente_nominal_a: 22.7
fator_potencia: 1.0
rendimento_pct: 97.5
dht_corrente_pct: 3"""

    colado_go = campo_colar_dados(exemplo_go, "[modulos], [inversores]")

    if colado_go:
        st.success("Texto reconhecido! Confira o resumo abaixo antes de gerar.")
        blocos = colado_go.pop("_blocos", {})
        modulos = blocos.get("modulos", [])
        inversores = blocos.get("inversores", [])

        faltando = [c for c in ("nome", "cpf_cnpj") if not colado_go.get(c)]
        st.write(
            f"**{colado_go.get('nome','?')}** — CPF/CNPJ {colado_go.get('cpf_cnpj','?')} — "
            f"{len(modulos)} modelo(s) de módulo, {len(inversores)} modelo(s) de inversor"
        )
        with st.expander("Ver dados completos reconhecidos"):
            st.json({**colado_go, "modulos": modulos, "inversores": inversores})

        if faltando:
            st.error(f"Faltam campos obrigatórios no texto colado: {', '.join(faltando)}")
            return

        foto_go = st.file_uploader(
            "Print do mapa / foto de satélite com a marcação do imóvel (opcional)",
            type=["png", "jpg", "jpeg"], key="foto_colado_go",
        )

        if st.button("Gerar documentos a partir do texto colado (Equatorial-GO)", type="primary"):
            uf_go = colado_go.get("uf", "GO")
            data_str = colado_go.get("data_inicio_operacao")
            data_inicio = None
            if data_str:
                for fmt in ("%b %d, %Y", "%B %d, %Y", "%d/%m/%Y", "%Y-%m-%d"):
                    try:
                        data_inicio = datetime.strptime(str(data_str).strip(), fmt)
                        break
                    except ValueError:
                        continue
            if not data_inicio:
                data_inicio = datetime.today() + timedelta(days=15)
            endereco_go = colado_go.get("endereco", "")
            bairro_go = colado_go.get("bairro", "")

            dados = {
                "responsavel_tecnico": {
                    "nome": "Alex Gomes da Silva", "titulo_profissional": "Engenheiro de Energia",
                    "registro_profissional": 436143, "uf_registro": "MT",
                    "email": "alexgomes.eng.energ@gmail.com", "telefone": "66999645914",
                },
                "premissas_fixas": {
                    "tarifa_branca": "NÃO", "vistoria_apos_solicitacao": "NÃO",
                    "autoriza_entrega_contratos": "SIM", "declara_conformidade_normas": "SIM",
                    "grid_zero": "NÃO", "gratuidade_ren": "NÃO",
                    "autoriza_faturas_email": "NÃO", "declara_veracidade": "SIM",
                },
                "_referencia_hakknner": REFERENCIA_HAKKNNER,
                "cliente": {
                    "nome": colado_go.get("nome", ""), "cpf_cnpj": colado_go.get("cpf_cnpj", ""),
                    "celular": colado_go.get("celular", ""), "telefone_fixo": "",
                    "endereco": endereco_go, "email": colado_go.get("email", ""),
                    "cep": colado_go.get("cep", ""), "municipio": colado_go.get("municipio", "Aragarças"),
                    "uf": uf_go, "uf_extenso": "Goiás" if uf_go == "GO" else uf_go,
                    "receber_fatura_email": "NÃO",
                    "tipo_orcamento": "Orçamento de Conexão",
                    "uc_existente": colado_go.get("uc_existente", ""),
                    "tipo_solicitacao": "CONEXÃO DE GD EM UNIDADE CONSUMIDORA EXISTENTE SEM AUMENTO DE POTÊNCIA DISPONIBILIZADA (ver item abaixo)",
                    "cargas_especiais": "NÃO",
                    "ramo_atividade": colado_go.get("classe", "Residencial"),
                    "classe": colado_go.get("classe", "Residencial"),
                    "tipo_ligacao": colado_go.get("tipo_ligacao", "MONOFÁSICO"),
                    "tensao_atendimento_v": colado_go.get("tensao_atendimento_v", 220),
                    "carga_declarada_kw": colado_go.get("carga_declarada_kw", 10),
                    "disjuntor_entrada_a": colado_go.get("disjuntor_entrada_a", 50),
                    "potencia_disponibilizada_kw": colado_go.get("potencia_disponibilizada_kw", 10),
                    "tipo_ramal": colado_go.get("tipo_ramal", "AÉREO"),
                    "fuso": colado_go.get("fuso", ""), "coord_x": colado_go.get("coord_x", ""),
                    "coord_y": colado_go.get("coord_y", ""),
                    "modalidade_compensacao": colado_go.get("modalidade_compensacao", "AUTOCONSUMO LOCAL"),
                    "data_inicio_operacao": data_inicio.date().isoformat(),
                    "conta_contrato": colado_go.get("uc_existente", ""), "bairro": bairro_go,
                    "mes_ano_documento": f"{MES_PT[data_inicio.month-1].capitalize()} – {data_inicio.year}",
                    "telefone_cliente": colado_go.get("celular", ""),
                    "data_conclusao_geradora": data_inicio.strftime("%d/%m/%Y"),
                    "endereco_completo_comissionamento": f"{endereco_go.upper()}\n{bairro_go.upper()}",
                    "modulos": modulos, "inversores": inversores,
                },
            }
            for aviso in avisos_campos_atencao(dados["cliente"], CAMPOS_ATENCAO_EQUATORIAL):
                st.warning(f"⚠️ {aviso}")
            gerar_equatorial(dados, foto_go)
        return

    st.divider()
    if defaults:
        st.caption("💡 Os campos abaixo já vêm preenchidos com os dados confirmados no "
                   "quadro do Drive acima — revise mais uma vez antes de gerar.")
    resp = responsavel_tecnico_form()
    st.caption("Registro Profissional (Anexo I): 436143 · Registro Memorial: 1217762256 (fixos)")

    st.subheader("1. Identificação")
    c1, c2, c3 = st.columns(3)
    nome = c1.text_input("Nome completo", defaults.get("nome", ""))
    cpf_cnpj = c2.text_input("CPF/CNPJ", defaults.get("cpf_cnpj", ""))
    celular = c3.text_input("Celular", defaults.get("celular", ""))

    c1, c2 = st.columns(2)
    endereco = c1.text_input("Endereço (rua, número, complemento)", defaults.get("endereco", ""))
    email = c2.text_input("E-mail", defaults.get("email", ""))

    c1, c2, c3, c4 = st.columns(4)
    cep = c1.text_input("CEP", _txt(defaults, "cep", "76240000"))
    municipio = c2.text_input("Município", _txt(defaults, "municipio", "Aragarças"))
    uf = c3.text_input("UF", _txt(defaults, "uf", "GO"))
    bairro = c4.text_input("Bairro", defaults.get("bairro", ""))

    uc_existente = st.text_input("UC (se já existir)", defaults.get("uc_existente", ""))

    st.subheader("2. Características técnicas")
    c1, c2, c3 = st.columns(3)
    opcoes_ligacao = ["MONOFÁSICO", "BIFÁSICO", "TRIFÁSICO"]
    tipo_ligacao = c1.selectbox("Tipo de Ligação", opcoes_ligacao,
                                 index=_idx(opcoes_ligacao, defaults.get("tipo_ligacao")))
    tensao_atendimento_v = c2.number_input("Tensão de Atendimento (V)",
                                            value=_num(defaults, "tensao_atendimento_v", 220))
    opcoes_classe = ["Residencial", "Comercial", "Industrial", "Rural"]
    classe = c3.selectbox("Classe", opcoes_classe, index=_idx(opcoes_classe, defaults.get("classe")))

    c1, c2, c3 = st.columns(3)
    disjuntor_entrada_a = c1.number_input("Disjuntor de Entrada (A)", value=_num(defaults, "disjuntor_entrada_a", 50))
    carga_declarada_kw = c2.number_input("Carga Declarada (kW)", value=_num(defaults, "carga_declarada_kw", 10.0))
    potencia_disponibilizada_kw = c3.number_input("Potência Disponibilizada — PD (kW)",
                                                    value=_num(defaults, "potencia_disponibilizada_kw", 10.0))

    opcoes_ramal_go = ["AÉREO", "SUBTERRÂNEO"]
    tipo_ramal = st.selectbox("Tipo de Ramal", opcoes_ramal_go,
                               index=_idx(opcoes_ramal_go, defaults.get("tipo_ramal")))

    st.subheader("3. Coordenadas e modalidade")
    c1, c2, c3 = st.columns(3)
    fuso = c1.text_input("Fuso (ex: 22L)", defaults.get("fuso", ""))
    coord_x = c2.text_input("X", defaults.get("coord_x", ""))
    coord_y = c3.text_input("Y", defaults.get("coord_y", ""))

    c1, c2 = st.columns(2)
    opcoes_modalidade_go = ["AUTOCONSUMO LOCAL", "AUTOCONSUMO REMOTO"]
    modalidade_compensacao = c1.selectbox("Modalidade de Compensação", opcoes_modalidade_go,
                                           index=_idx(opcoes_modalidade_go, defaults.get("modalidade_compensacao")))
    data_inicio = c2.date_input("Data de início de operação", value=date.today() + timedelta(days=15))

    st.subheader("4. Módulos")
    caixa_fonte_inmetro()
    modulos = tabela_equipamentos(
        "Um item por modelo de módulo diferente",
        None,
        [{"potencia_w": 0, "quantidade": 0, "fabricante": "", "modelo": ""}],
        "modulos_equatorial",
    )

    st.subheader("5. Inversor(es)")
    st.caption("Busque o datasheet real do fabricante para corrente/rendimento — não estime.")
    inversores = tabela_equipamentos(
        "Um item por modelo de inversor diferente",
        None,
        [{"fabricante": "", "modelo": "", "potencia_nominal_kw": 0.0, "faixa_tensao_v": 220,
          "corrente_nominal_a": 0.0, "fator_potencia": 1.0, "rendimento_pct": 98.0, "dht_corrente_pct": 3.0}],
        "inversores_equatorial",
    )

    st.subheader("6. Foto de localização (opcional)")
    foto = st.file_uploader("Print do mapa / foto de satélite com a marcação do imóvel", type=["png", "jpg", "jpeg"])

    gerar = st.button("Gerar documentos (Equatorial-GO)", type="primary")

    if gerar:
        if not nome or not cpf_cnpj:
            st.error("Preencha ao menos Nome e CPF/CNPJ.")
            return

        dados = {
            "responsavel_tecnico": {
                "nome": resp["nome"], "titulo_profissional": "Engenheiro de Energia",
                "registro_profissional": 436143, "uf_registro": "MT",
                "email": resp["email"], "telefone": resp["telefone"].replace(" ", "").replace("(", "").replace(")", "").replace("-", ""),
            },
            "premissas_fixas": {
                "tarifa_branca": "NÃO", "vistoria_apos_solicitacao": "NÃO",
                "autoriza_entrega_contratos": "SIM", "declara_conformidade_normas": "SIM",
                "grid_zero": "NÃO", "gratuidade_ren": "NÃO",
                "autoriza_faturas_email": "NÃO", "declara_veracidade": "SIM",
            },
            "_referencia_hakknner": REFERENCIA_HAKKNNER,
            "cliente": {
                "nome": nome, "cpf_cnpj": cpf_cnpj, "celular": celular, "telefone_fixo": "",
                "endereco": endereco, "email": email, "cep": cep, "municipio": municipio,
                "uf": uf, "uf_extenso": "Goiás" if uf == "GO" else uf,
                "receber_fatura_email": "NÃO",
                "tipo_orcamento": "Orçamento de Conexão", "uc_existente": uc_existente,
                "tipo_solicitacao": "CONEXÃO DE GD EM UNIDADE CONSUMIDORA EXISTENTE SEM AUMENTO DE POTÊNCIA DISPONIBILIZADA (ver item abaixo)",
                "cargas_especiais": "NÃO", "ramo_atividade": classe, "classe": classe,
                "tipo_ligacao": tipo_ligacao, "tensao_atendimento_v": tensao_atendimento_v,
                "carga_declarada_kw": carga_declarada_kw, "disjuntor_entrada_a": disjuntor_entrada_a,
                "potencia_disponibilizada_kw": potencia_disponibilizada_kw, "tipo_ramal": tipo_ramal,
                "fuso": fuso, "coord_x": coord_x, "coord_y": coord_y,
                "modalidade_compensacao": modalidade_compensacao,
                "data_inicio_operacao": data_inicio.isoformat(),
                "conta_contrato": uc_existente, "bairro": bairro,
                "mes_ano_documento": f"{MES_PT[data_inicio.month-1].capitalize()} – {data_inicio.year}",
                "telefone_cliente": celular,
                "data_conclusao_geradora": data_inicio.strftime("%d/%m/%Y"),
                "endereco_completo_comissionamento": f"{endereco.upper()}\n{bairro.upper()}",
                "modulos": modulos, "inversores": inversores,
            },
        }
        for aviso in avisos_campos_atencao(dados["cliente"], CAMPOS_ATENCAO_EQUATORIAL):
            st.warning(f"⚠️ {aviso}")
        gerar_equatorial(dados, foto)


def gerar_equatorial(dados, foto_upload):
    nome_base = slug_arquivo(dados["cliente"]["nome"], dados["cliente"]["uc_existente"] or "SEMUC")
    with tempfile.TemporaryDirectory() as tmp:
        tmp = Path(tmp)

        with st.spinner("Preenchendo Anexo I..."):
            xlsx_path = tmp / f"{nome_base}_AnexoI.xlsx"
            avisos = preencher_anexo1(dados, str(MODELO_ANEXO1), str(xlsx_path))
            resultado = recalc_xlsx(str(xlsx_path), 90)

        with st.spinner("Gerando Memorial Descritivo..."):
            doc = Document(str(MODELO_MEMORIAL))
            replace_everywhere(doc, montar_substituicoes_memorial(dados))
            qtd = sum(m["quantidade"] for m in dados["cliente"]["modulos"])
            pot = sum(m["potencia_w"] * m["quantidade"] for m in dados["cliente"]["modulos"]) / 1000
            fix_tabela_gerador(doc, qtd, pot)
            if foto_upload is not None:
                foto_path = tmp / "foto_local.png"
                foto_path.write_bytes(foto_upload.getvalue())
                substituir_foto_localizacao(doc, str(foto_path))
            memorial_path = tmp / f"{nome_base}_Memorial.docx"
            doc.save(str(memorial_path))

        with st.spinner("Gerando Relatório de Comissionamento..."):
            doc2 = Document(str(MODELO_COMISSIONAMENTO))
            replace_everywhere(doc2, montar_substituicoes_comissionamento(dados))
            comiss_path = tmp / f"{nome_base}_Comissionamento.docx"
            doc2.save(str(comiss_path))

        with st.spinner("Convertendo para PDF..."):
            run_soffice(["--headless", "--convert-to", "pdf", "--outdir", str(tmp), str(xlsx_path)])
            run_soffice(["--headless", "--convert-to", "pdf", "--outdir", str(tmp), str(memorial_path)])
            run_soffice(["--headless", "--convert-to", "pdf", "--outdir", str(tmp), str(comiss_path)])

            from pypdf import PdfReader, PdfWriter
            anexo_pdf_full = xlsx_path.with_suffix(".pdf")
            reader = PdfReader(str(anexo_pdf_full))
            writer = PdfWriter()
            for i in [1, 2, 3, 4]:
                if i < len(reader.pages):
                    writer.add_page(reader.pages[i])
            anexo_pdf = tmp / f"{nome_base}_AnexoI.pdf"
            with open(anexo_pdf, "wb") as f:
                writer.write(f)

        # Guarda os bytes na sessão (a pasta temporária é apagada ao sair
        # deste bloco, e clicar num botão de download recarrega a página).
        st.session_state["saida_equatorial"] = {
            "avisos": avisos,
            "anexo_xlsx_bytes": xlsx_path.read_bytes(), "anexo_xlsx_nome": xlsx_path.name,
            "anexo_pdf_bytes": anexo_pdf.read_bytes(), "anexo_pdf_nome": anexo_pdf.name,
            "memorial_docx_bytes": memorial_path.read_bytes(), "memorial_docx_nome": memorial_path.name,
            "memorial_pdf_bytes": memorial_path.with_suffix(".pdf").read_bytes(),
            "memorial_pdf_nome": memorial_path.with_suffix(".pdf").name,
            "comiss_docx_bytes": comiss_path.read_bytes(), "comiss_docx_nome": comiss_path.name,
            "comiss_pdf_bytes": comiss_path.with_suffix(".pdf").read_bytes(),
            "comiss_pdf_nome": comiss_path.with_suffix(".pdf").name,
        }

    mostrar_downloads_equatorial()


def mostrar_downloads_equatorial():
    saida = st.session_state.get("saida_equatorial")
    if not saida:
        return
    if saida["avisos"]:
        st.warning("⚠️ Checagem de viabilidade do Anexo I:\n" + "\n".join(f"- {a}" for a in saida["avisos"]))
    else:
        st.success("Checagem de viabilidade: OK")

    st.success("Documentos gerados!")
    c1, c2, c3, c4, c5, c6, c7 = st.columns(7)
    c1.download_button("⬇️ Anexo I (.xlsx)", saida["anexo_xlsx_bytes"], saida["anexo_xlsx_nome"], key="dl_anexo_xlsx")
    c2.download_button("⬇️ Anexo I (.pdf)", saida["anexo_pdf_bytes"], saida["anexo_pdf_nome"], key="dl_anexo_pdf")
    c3.download_button("⬇️ Memorial (.docx)", saida["memorial_docx_bytes"], saida["memorial_docx_nome"], key="dl_mem_docx")
    c4.download_button("⬇️ Memorial (.pdf)", saida["memorial_pdf_bytes"], saida["memorial_pdf_nome"], key="dl_mem_pdf")
    c5.download_button("⬇️ Comissionamento (.docx)", saida["comiss_docx_bytes"], saida["comiss_docx_nome"], key="dl_com_docx")
    c6.download_button("⬇️ Comissionamento (.pdf)", saida["comiss_pdf_bytes"], saida["comiss_pdf_nome"], key="dl_com_pdf")
    zip_bytes = zip_de_arquivos([
        (saida["anexo_xlsx_nome"], saida["anexo_xlsx_bytes"]),
        (saida["anexo_pdf_nome"], saida["anexo_pdf_bytes"]),
        (saida["memorial_docx_nome"], saida["memorial_docx_bytes"]),
        (saida["memorial_pdf_nome"], saida["memorial_pdf_bytes"]),
        (saida["comiss_docx_nome"], saida["comiss_docx_bytes"]),
        (saida["comiss_pdf_nome"], saida["comiss_pdf_bytes"]),
    ])
    c7.download_button("📦 Baixar tudo (.zip)", zip_bytes, "documentos_equatorial.zip",
                        key="dl_zip_equatorial", type="primary")
    st.info("Lembrete: as seções de dimensionamento do Memorial (disjuntor, DPS, "
            "aterramento, cabos, levantamento de carga) usam os valores de referência — revise manualmente.")


# ============================================================
# MAIN
# ============================================================
if LOGO_PATH.exists():
    col_logo, col_titulo = st.columns([1, 5], vertical_alignment="center")
    col_logo.image(str(LOGO_PATH), width=110)
    col_titulo.title("Gerador de Documentos GD")
else:
    st.title("☀️ Gerador de Documentos GD — Exergia")

concessionaria = st.sidebar.radio("Concessionária", ["Energisa-MT", "Equatorial-GO"])
st.sidebar.markdown("---")
st.sidebar.caption("Preenchimento automático de projetos de Geração Distribuída, "
                    "com as regras e validações já testadas no fluxo de produção da Exergia.")

if concessionaria == "Energisa-MT":
    form_energisa()
else:
    form_equatorial()
