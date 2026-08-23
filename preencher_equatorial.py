#!/usr/bin/env python3
"""
Preenchimento automático do ANEXO I - Formulário de Solicitação de Orçamento
de Microgeração Distribuída Grupo B (Equatorial Goiás - NT.00020.EQTL).

Uso:
    python3 preencher_equatorial.py dados_cliente.json modelo_anexo_i.xlsx saida.xlsx

Ao final, verifica automaticamente as 3 células de viabilidade do projeto
(C33, Y61) e avisa se o projeto não é viável como especificado.
"""
import sys
import json
import shutil
import datetime
import openpyxl

MAX_MODULOS = 10   # GUIA 0, linhas 7-16
MAX_INVERSORES = 30  # GUIA 0, linhas 22-51


def _parse_data(valor):
    """Aceita 'AAAA-MM-DD' ou 'DD/MM/AAAA' e devolve um date real, para a
    célula guardar a data como valor de data do Excel (o modelo da
    Equatorial exibe esse campo em formato numérico serial, ex: 46.245,00)."""
    if isinstance(valor, (datetime.date, datetime.datetime)):
        return valor
    if "-" in valor:
        return datetime.datetime.strptime(valor, "%Y-%m-%d").date()
    return datetime.datetime.strptime(valor, "%d/%m/%Y").date()


def preencher(dados: dict, modelo_path: str, saida_path: str):
    shutil.copyfile(modelo_path, saida_path)
    wb = openpyxl.load_workbook(saida_path, data_only=False)

    cliente = dados["cliente"]
    resp_tecnico = dados["responsavel_tecnico"]
    resp_legal = dados.get("responsavel_legal", resp_tecnico)
    fixo = dados.get("premissas_fixas", {})

    # ---------------- GUIA "0" — equipamentos ----------------
    ws = wb["0"]
    for i, m in enumerate(cliente.get("modulos", [])[:MAX_MODULOS]):
        r = 7 + i
        ws[f"D{r}"] = m["potencia_w"]
        ws[f"H{r}"] = m["quantidade"]
        ws[f"T{r}"] = m["fabricante"]
        ws[f"AA{r}"] = m["modelo"]
        # K (potência de pico kWp) e P (área m²) são calculados automaticamente

    for i, inv in enumerate(cliente.get("inversores", [])[:MAX_INVERSORES]):
        r = 22 + i
        ws[f"D{r}"] = inv["fabricante"]
        ws[f"H{r}"] = inv["modelo"]
        ws[f"L{r}"] = inv["potencia_nominal_kw"]
        ws[f"P{r}"] = inv["faixa_tensao_v"]
        ws[f"T{r}"] = inv["corrente_nominal_a"]
        ws[f"W{r}"] = inv["fator_potencia"]
        ws[f"Z{r}"] = inv["rendimento_pct"]
        ws[f"AC{r}"] = inv["dht_corrente_pct"]

    # ---------------- GUIA "1" — dados cadastrais ----------------
    ws = wb["1"]
    ws["C10"] = cliente["nome"]
    ws["R10"] = cliente["cpf_cnpj"]
    ws["Y10"] = cliente["celular"]
    ws["AB10"] = cliente.get("telefone_fixo", "")
    ws["C13"] = cliente["endereco"]
    ws["R13"] = cliente["email"]
    ws["D15"] = cliente["cep"]
    ws["I15"] = cliente["municipio"]
    ws["Q15"] = cliente["uf"]
    ws["Z15"] = cliente.get("receber_fatura_email", "NÃO")

    ws["F17"] = cliente.get("tipo_orcamento", "Orçamento de Conexão")
    if cliente.get("uc_existente"):
        ws["Z17"] = cliente["uc_existente"]

    ws["G19"] = cliente["tipo_solicitacao"]

    ws["F23"] = fixo.get("tarifa_branca", "NÃO")
    ws["F25"] = cliente.get("cargas_especiais", "NÃO")
    if cliente.get("detalhar_cargas_especiais"):
        ws["H25"] = cliente["detalhar_cargas_especiais"]

    ws["G27"] = cliente["ramo_atividade"]
    ws["F29"] = cliente["classe"]
    ws["T29"] = cliente["tipo_ligacao"]
    ws["AC29"] = cliente["tensao_atendimento_v"]

    ws["F31"] = cliente["carga_declarada_kw"]
    ws["P31"] = cliente["disjuntor_entrada_a"]
    ws["AB31"] = cliente["potencia_disponibilizada_kw"]
    ws["P33"] = cliente["tipo_ramal"]

    ws["R35"] = cliente["fuso"]
    ws["T35"] = cliente["coord_x"]
    ws["Z35"] = cliente["coord_y"]

    ws["C38"] = resp_legal["nome"]
    ws["R38"] = resp_legal["telefone"]
    ws["Y38"] = resp_legal["email"]

    ws["C43"] = resp_tecnico["nome"]
    ws["M43"] = resp_tecnico["titulo_profissional"]
    ws["Y43"] = resp_tecnico["registro_profissional"]
    ws["AE43"] = resp_tecnico["uf_registro"]
    ws["C46"] = resp_tecnico["email"]
    ws["S46"] = resp_tecnico["telefone"]

    ws["G51"] = cliente.get("tipo_fonte_primaria", "SOLAR FOTOVOLTAICA")
    ws["G53"] = cliente.get("tipo_geracao", "EMPREGANDO CONVERSOR ELETRÔNICO/INVERSOR")
    ws["I55"] = cliente["modalidade_compensacao"]

    # PGT (kW) = min(soma dos módulos, soma dos inversores) — usado para decidir
    # automaticamente se o projeto se enquadra no "Fast Track" (limite 7,5 kW)
    pgt_paineis = sum(
        (m["potencia_w"] * m["quantidade"]) / 1000 for m in cliente.get("modulos", [])
    )
    pgt_inversores = sum(i["potencia_nominal_kw"] for i in cliente.get("inversores", []))
    pgt = min(pgt_paineis, pgt_inversores) if pgt_inversores else pgt_paineis
    fast_track_elegivel = pgt <= 7.5 and cliente["modalidade_compensacao"] == "AUTOCONSUMO LOCAL"
    data_op = _parse_data(cliente["data_inicio_operacao"])
    fmt_original = ws["U61"].number_format
    ws["U61"] = data_op
    ws["U61"].number_format = fmt_original

    # Solicitações e declarações (padrão fixo, salvo indicação contrária)
    ws["AD92"] = fixo.get("vistoria_apos_solicitacao", "NÃO")
    ws["AD93"] = fixo.get("autoriza_entrega_contratos", "SIM")
    ws["AD94"] = fixo.get("declara_conformidade_normas", "SIM")
    ws["AD96"] = fixo.get("grid_zero", "NÃO")
    ws["AD97"] = fixo.get("gratuidade_ren", "NÃO")
    ws["AD98"] = cliente.get("fast_track_forcar") or ("SIM" if fast_track_elegivel else "NÃO")
    ws["AD99"] = fixo.get("autoriza_faturas_email", "NÃO")
    ws["AD100"] = fixo.get("declara_veracidade", "SIM")

    wb.save(saida_path)

    # ---------------- Checagem de viabilidade ----------------
    wb_check = openpyxl.load_workbook(saida_path, data_only=False)
    return checar_viabilidade(wb_check)


def checar_viabilidade(wb):
    """Recalcula não é possível sem Excel/LibreOffice; aqui fazemos a checagem
    manualmente com a mesma lógica das fórmulas C33 e Y61, para avisar o
    usuário mesmo antes de abrir o arquivo."""
    ws = wb["1"]
    avisos = []

    carga = ws["F31"].value
    pd = ws["AB31"].value
    if carga is not None and pd is not None:
        if carga > pd:
            avisos.append(
                f"NOK: Carga Declarada ({carga} kW) > Potência Disponibilizada ({pd} kW)"
            )

    # PGT = min(soma das potências de pico dos módulos, potência dos inversores) — a
    # planilha limita a geração à capacidade AC dos inversores, não à soma DC dos painéis.
    ws0 = wb["0"]
    soma_paineis_kwp = 0
    for r in range(7, 17):
        v = ws0[f"D{r}"].value
        q = ws0[f"H{r}"].value
        if v and q:
            soma_paineis_kwp += (v * q) / 1000
    soma_inversores_kw = 0
    for r in range(22, 52):
        p = ws0[f"L{r}"].value
        if p:
            soma_inversores_kw += p
    pgt = min(soma_paineis_kwp, soma_inversores_kw) if soma_inversores_kw else soma_paineis_kwp
    if pd is not None:
        if pgt > 75:
            avisos.append(f"PGT ACIMA DO LIMITE DO GRUPO B (PGT={pgt:.2f} kW, limite 75 kW)")
        elif pgt > pd:
            avisos.append(f"NOK: PGT ({pgt:.2f} kW) > Potência Disponibilizada ({pd} kW)")

    fast_track = ws["AD98"].value
    if fast_track == "SIM" and pgt > 7.5:
        avisos.append(f"PGT Acima do Limite de 7,5 kW para Fast Track (PGT={pgt:.2f} kW)")

    return avisos


if __name__ == "__main__":
    if len(sys.argv) != 4:
        print("Uso: python3 preencher_equatorial.py dados_cliente.json modelo_anexo_i.xlsx saida.xlsx")
        sys.exit(1)
    with open(sys.argv[1], encoding="utf-8") as f:
        dados = json.load(f)
    avisos = preencher(dados, sys.argv[2], sys.argv[3])
    print(f"Planilha preenchida salva em: {sys.argv[3]}")
    if avisos:
        print("\n⚠️  ATENÇÃO — projeto pode não ser viável:")
        for a in avisos:
            print("  -", a)
    else:
        print("Checagem de viabilidade: OK.")
